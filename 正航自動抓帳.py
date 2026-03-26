from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
import re
import gc
import win32com.client as win32
import matplotlib.pyplot as plt
from pandas.plotting import table
import pyautogui
import time
import subprocess
import pyperclip
import shutil
import json




file_name_1 = "會計拋轉需要_明細表_"
file_name_2 = "會計拋轉需要_簡要表_"

grouped_data = []
current_date = None
current_sum = 0
opening_balance_sum = 0  # ✅ 新增：庫存期初 / 無日期


#明細表的位置
file_name_path = r"\\192.168.2.253\\管理中心\\財會課\\加盟店應收\\應收拋轉(心豪)"
#明細表日期是簡要表減1天工作天
date1_start ="20250607"
date1_end = "20260325"
#簡要表日期2個工作天
date2_start ="20250607"
date2_end = "20260325"
# 判斷星期幾
dt = datetime.strptime(date2_end, "%Y%m%d")
weekday_index = dt.weekday()

base_path = r"C:\\Users\\楊心豪\\Desktop\\爬蟲"
store_list_json = r"\\192.168.2.253\\事業中心\\資訊\\code\\store_list.json"
# 你的 NAS 路徑（請依實際情況修改）
nas_base_path = r"\\192.168.2.253\\管理中心\\財會課\\加盟店應收\\應收拋轉(心豪)"
output_folder = os.path.join(base_path, date2_end)
# 確保資料夾存在（不存在就建立）
os.makedirs(output_folder, exist_ok=True)
# 1. 讀入完整代號 vs 門市名稱 Excel
#df_all = pd.read_excel(r"C:\\Users\\楊心豪\Desktop\\爬蟲\\TEATOP店家通訊錄(最後更新2025.3.3).xlsx")

from datetime import datetime, timedelta

def subtract_workday(date_str, days=1):
    dt = datetime.strptime(date_str, "%Y%m%d")
    while days > 0:
        dt -= timedelta(days=1)
        if dt.weekday() < 5:  # 0~4 = 工作天
            days -= 1
    return dt

# 簡要表實際日期
summary_dt = datetime.strptime(date2_end, "%Y%m%d")

# 明細表實際日期 = 簡要表減 1 個工作天
detail_dt = subtract_workday(date2_end, 1)

# ======================
# 星期判斷（只算一次）
# ======================
weekday_dict = {
    0: "monday",
    1: "tuesday",
    2: "wednesday",
    3: "thursday",
    4: "friday",
    5: "saturday",
    6: "sunday",
}

weekday_index = summary_dt.weekday()
weekday_name = weekday_dict[weekday_index]

print("📅 簡要表日期：", summary_dt.strftime("%Y-%m-%d"))
print("📅 明細表日期：", detail_dt.strftime("%Y-%m-%d"))
print("📅 使用星期：", weekday_name)

# ======================
# 檔案路徑
# ======================
output_folder = os.path.join(base_path, date2_end)
os.makedirs(output_folder, exist_ok=True)

file_name_full = file_name_2 + date2_end
source_file = os.path.join(
    file_name_path,
    file_name_full + ".xlsx"
)

cleaned_file = os.path.join(
    r"C:\\Users\\楊心豪\\Downloads",
    "已清理" + file_name_full + ".xlsx"
)

# ======================
# 清理簡要表（刪第一列）
# ======================
sheet_names = pd.ExcelFile(source_file).sheet_names

with pd.ExcelWriter(cleaned_file, engine="openpyxl") as writer:
    for sheet in sheet_names:
        df = pd.read_excel(source_file, sheet_name=sheet, header=None)
        df = df.iloc[1:].reset_index(drop=True)
        df.to_excel(writer, sheet_name=sheet, index=False, header=False)

print("✅ 所有工作表的第一列已刪除，儲存為：", cleaned_file)

source_file = cleaned_file

# ======================
# 讀 store_list.json
# ======================
with open(store_list_json, "r", encoding="utf-8") as f:
    store_list = json.load(f)

# ======================
# 建立 weekday → {代號: 店名}
# ======================
dicts_by_week = {}

for store_name, info in store_list.items():
    weekday = info.get("weekday")
    code = str(info.get("code", "")).strip().upper()
    if not weekday or not code:
        continue
    dicts_by_week.setdefault(weekday, {})[code] = store_name

# ======================
# 取得當天要用的門市字典
# ======================
selected_dict = dicts_by_week.get(weekday_name, {})

print(f"📦 {weekday_name} 門市數量：", len(selected_dict))

# ======================
# 載入 Excel，準備後續處理
# ======================
wb = load_workbook(source_file)

def get_customer_id(ws):
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=5):
        for cell in row:
            val = str(cell.value) if cell.value else ""
            if "客戶編號" in val:
                print(f"📌 找到原始內容：{val}")
                if "：" in val:
                    return val.split("：")[1].strip().upper()
    return ""


import os
import pandas as pd
from openpyxl import load_workbook



# 儲存符合條件的分頁資料
for sheetname in wb.sheetnames:
    ws = wb[sheetname]

    customer_id = get_customer_id(ws)
    store_name = str(ws["C6"].value).strip() if ws["C6"].value else ""
    if customer_id in selected_dict:
        file_name = f"{store_name}_簡要表_{date2_end}.xlsx"
        full_path = os.path.join(output_folder, file_name)

        # 不使用第一列當標題
        df = pd.read_excel(source_file, sheet_name=sheetname, engine="openpyxl", header=None)

        # 刪除第一列（原始格式或空白列）
        df = df.iloc[1:].reset_index(drop=True)

        # 儲存檔案（不加 index）
        df.to_excel(full_path, index=False, header=False)

        print(f"已儲存：{full_path}")
    else:
        print(f"分頁 {sheetname} 的客戶編號 {customer_id} 不在選擇字典中")

# -------- 開始讀取儲存後的檔案，抽取資料 --------
import re

result_list = []

for filename in os.listdir(output_folder):
    if filename.endswith(".xlsx"):
        filepath = os.path.join(output_folder, filename)
        df = pd.read_excel(filepath, header=None, engine="openpyxl")

        store_name = None
        date_row_idx = None
        date_col_idx = None
        unpaid_col_idx = None
        prepayment = 0  # 預收金額初始化

        # 找出公司名稱
        for i in range(df.shape[0]):
            row = df.iloc[i]
            for j in range(len(row)):
                if str(row[j]).strip() == "公司名稱：":
                    for k in range(j + 1, len(row)):
                        val = str(row[k]).strip()
                        if val and val != "0" and val.lower() != "nan":
                            store_name = val
                            break

        # 找出「日期」和「未收金額」欄位位置
        for i in range(df.shape[0]):
            row = df.iloc[i]
            for j in range(len(row)):
                if "日期" in str(row[j]):
                    date_row_idx = i
                    date_col_idx = j
                if "未收金額" in str(row[j]):
                    unpaid_col_idx = j
            if date_row_idx is not None and unpaid_col_idx is not None:
                break

        if date_row_idx is None or unpaid_col_idx is None:
            print(f"⚠️ 檔案 {filename} 找不到『日期』或『未收金額』欄位，跳過")
            continue

        # 初始化變數
        grouped_data = []  # 存每個日期區段的加總
        current_date = None
        current_sum = 0

        for i in range(date_row_idx + 1, df.shape[0]):
            row = df.iloc[i]

            # 如果遇到「本幣累計預收貨款金額」 → 用正則抓第一個數字
            for j in range(len(row)):
                if "本幣累計預收貨款金額" in str(row[j]):
                    row_str = " ".join(map(str, row))  # 整行轉字串
                    match = re.search(r"\d+", row_str)
                    if match:
                        prepayment = float(match.group())
                    else:
                        prepayment = 0
                    break
            if "本幣累計預收貨款金額" in " ".join(map(str, row)):
                break

            raw_date = str(row[date_col_idx]).strip()
            unpaid_val = row[unpaid_col_idx]

            parsed_date = pd.to_datetime(raw_date, errors='coerce')
            is_new_date = pd.notna(parsed_date)

            if is_new_date:
                # 遇到新日期，先結算上一段
                if current_date is not None:
                    grouped_data.append((current_date, current_sum))
                current_date = parsed_date
                current_sum = 0

            # 加總未收金額
            if pd.notna(unpaid_val):
                try:
                    val = float(str(unpaid_val).replace(",", "").strip())

                    if current_date is None:
                        # ✅ 還沒遇到任何日期 → 當作庫存期初 / 上期
                        opening_balance_sum += val
                    else:
                        current_sum += val

                except:
                    pass


        # 最後一段也加進去
        if current_date is not None:
            grouped_data.append((current_date, current_sum))

        # 拆分本期 / 上期（檢查是否有 date2_end）
        target_date = pd.to_datetime(date2_end, format="%Y%m%d", errors="coerce")
        has_current_period = any(d == target_date for d, _ in grouped_data)

        if grouped_data:
            if has_current_period:
                # 本期金額 = date2_end 當天的加總
                current_period_amount = sum(val for d, val in grouped_data if d == target_date)
                previous_period_amount = (
                    opening_balance_sum +
                    sum(val for d, val in grouped_data if d != target_date)
                )

            else:
                # 沒有當天叫貨 → 全部算上期
                current_period_amount = 0
                previous_period_amount = sum(val for _, val in grouped_data)
            total_receivable = current_period_amount + previous_period_amount
        else:
            current_period_amount = 0
            previous_period_amount = 0
            total_receivable = 0

        adjusted_total = total_receivable - (prepayment if prepayment else 0)  # ✅ 預收 NaN → 0

        if store_name:
            result_list.append({
                "店名": store_name,
                "上期": previous_period_amount,
                "本期": current_period_amount,
                "預收": prepayment if prepayment else 0,  # ✅ NaN → 0
                "本幣期末應收帳款總計(本期+上期-預收)": adjusted_total
            })
        else:
            print(f"❗檔案 {filename} 缺少資訊：{store_name=}, {total_receivable=}, {prepayment=}")

# 匯出彙總表
summary_df = pd.DataFrame(result_list)

# 再次確保「預收」為數字且 NaN → 0
summary_df["預收"] = pd.to_numeric(summary_df["預收"], errors="coerce").fillna(0)
summary_df["本幣期末應收帳款總計(本期+上期-預收)"] = (
    summary_df["上期"] + summary_df["本期"] - summary_df["預收"]
)

summary_df.to_excel(os.path.join(output_folder, "_應收彙總.xlsx"), index=False)
print(summary_df)


time.sleep(2)
# 啟動 Excel
excel = win32.gencache.EnsureDispatch("Excel.Application")
excel.Visible = False  # 不顯示 Excel 介面

for filename in os.listdir(output_folder):
    if filename.endswith(".xlsx") and not filename.endswith("_應收彙總.xlsx"):
        filepath = os.path.join(output_folder, filename)
        pdfpath = filepath.replace(".xlsx", ".pdf")

        try:
            wb = excel.Workbooks.Open(filepath)
            ws = wb.Sheets(1)  # 假設每個檔案都只有一個 sheet

            # 🧠 找到最後一個有內容的儲存格（自動推算範圍）
            last_row = ws.UsedRange.Rows.Count
            last_col = ws.UsedRange.Columns.Count

            # 🖨️ 設定列印區域
            ws.PageSetup.PrintArea = ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col)).Address

            ws.UsedRange.Font.Bold = True
            ws.UsedRange.Font.Size = 12  # 可選，加大字體（10~12 都很舒服）

            # ⚙️ 設定縮放：一頁寬、高不限
            ws.PageSetup.Orientation = 2 
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False

            # 匯出成 PDF
            wb.ExportAsFixedFormat(0, pdfpath)
            print(f"✅ 已轉換為 PDF：{pdfpath}")

            wb.Close(False)

        except Exception as e:
            print(f"❌ {filename} 發生錯誤：{e}")


f# 關閉 Excel
excel.Quit()

# 🔥 刪除除了 "_應收彙總.xlsx" 以外的所有 xlsx
for filename in os.listdir(output_folder):
    if filename.endswith(".xlsx") and not filename.endswith("_應收彙總.xlsx"):
        try:
            os.remove(os.path.join(output_folder, filename))
            print(f"🗑️ 已刪除：{filename}")
        except Exception as e:
            print(f"❌ 刪除 {filename} 發生錯誤：{e}")



# 🧾 來源 Excel：會計拋轉需要_明細表
source_file = os.path.join(file_name_path, f"會計拋轉需要_明細表_{date1_end}.xlsx")

# 📁 輸出資料夾（與簡要表同一個）
output_folder = os.path.join(base_path, date2_end)
os.makedirs(output_folder, exist_ok=True)

# 📃 取得簡要表的名單
summary_store_names = set()
for filename in os.listdir(output_folder):
    if filename.endswith("_應收彙總.xlsx"):
        df = pd.read_excel(os.path.join(output_folder, filename), engine="openpyxl")
        summary_store_names.update(df["店名"].astype(str).tolist())

time.sleep(5)

# ⚙️ 開啟 Excel 應用程式
excel = win32.gencache.EnsureDispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open(source_file)
excel = win32.gencache.EnsureDispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open(source_file)
time.sleep(5)
# 🔍 建立：公司名稱 → 對應工作表名清單
company_sheets = {}

for sheet in wb.Sheets:
    df = pd.read_excel(source_file, sheet_name=sheet.Name, engine="openpyxl", header=None)

    company_name = None
    for i in range(min(20, df.shape[0])):
        row = df.iloc[i]
        for val in row:
            if isinstance(val, str) and "公司名稱" in val:
                match = re.search(r"公司名稱[:：\s]*(\S+)", val)
                if match:
                    company_name = match.group(1).strip()
                break
        if company_name:
            break

    if company_name:
        company_sheets.setdefault(company_name, []).append(sheet.Name)

# 📤 輸出 PDF（只針對簡要表名單中的公司）
for company, sheets in company_sheets.items():
    if company not in summary_store_names:
        continue

    try:
        temp_wb = excel.Workbooks.Add()
        default_sheet = temp_wb.Sheets(1)
        copied = False

        # 🔁 關鍵：反向複製，讓 PDF 順序變成 1 → 2 → 3 → 4
        for sheet_name in reversed(sheets):
            sheet = wb.Sheets(sheet_name)
            sheet.Copy(Before=temp_wb.Sheets(1))
            copied = True

        if copied:
            default_sheet.Delete()

            for ws in temp_wb.Sheets:
                # 🧹 刪除前三列
                ws.Rows("1:3").Delete()
                # ✨ 新增：強制統一字體，避免 PDF 轉檔亂碼
                try:
                    ws.Cells.Font.Name = "微軟正黑體"
                    # 如果微軟正黑體抓不到，可以改用 "新細明體"
                except:
                    pass
                # 📐 設定列印區域
                last_row = ws.UsedRange.Rows.Count
                last_col = ws.UsedRange.Columns.Count
                ws.PageSetup.PrintArea = ws.Range(
                    ws.Cells(1, 1),
                    ws.Cells(last_row, last_col)
                ).Address

                # 📏 版面設定（完全不動你原本的比例）
                ws.PageSetup.Zoom = 70
                ws.PageSetup.Orientation = 2  # 橫向
                ws.PageSetup.CenterHorizontally = True
                ws.PageSetup.CenterVertically = True
            # 將 Quality 設為 0 (標準品質)，並確保 OpenAfterPublish 為 False
            temp_wb.ExportAsFixedFormat(0, pdf_path)
            safe_name = re.sub(r'[\\/*?:"<>|]', "_", company)
            pdf_path = os.path.join(
                output_folder,
                f"{safe_name}_明細表_{date1_end}.pdf"
            )

            temp_wb.ExportAsFixedFormat(0, pdf_path)
            print(f"✅ 已輸出 PDF：{pdf_path}")

        temp_wb.Close(False)

    except Exception as e:
        print(f"❌ {company} 錯誤：{e}")


# 🔚 關閉 Excel
wb.Close(False)
excel.Quit()
del excel
gc.collect()







# 組合完整目標路徑（例如：\\NAS_SERVER\共享資料\帳款報表\20240805）
nas_target_path = os.path.join(nas_base_path, date2_end)

try:
    # 如果目的資料夾已存在，先刪除（避免 shutil.move 報錯）
    if os.path.exists(nas_target_path):
        shutil.rmtree(nas_target_path)

    # 搬移整個資料夾
    shutil.move(output_folder, nas_target_path)
    print(f"📂 資料夾已成功搬移到 NAS：{nas_target_path}")

except Exception as e:
    print(f"❌ 搬移至 NAS 失敗：{e}")   



